# ==============================================================
# File: main.py
# Description:
#   Main automation script for RF testing across a frequency sweep.
#   Performs:
#     - Instrument setup (VSG, VSA, Power Meter, Servo)
#     - Baseline measurements (power, EVM, ACLR)
#     - Polynomial DPD, Direct DPD, GMP DPD correction loops (selectively enabled)
#     - Power Servo (external NRX) and K18 Power Servo (VSA-based)
#     - Results logging and statistics generation in Excel
# ==============================================================

import os
import json
import pandas as pd
import numpy as np
import logging
from time import time
from openpyxl.styles import Font, PatternFill, Alignment
from src.measurements.vsg import VSG
from src.measurements.vsa import VSA
from src.measurements.power_meter import PowerMeter
from src.measurements.power_servo import PowerServo
from src.measurements.et import ET  # Import ET class for envelope tracking

# --------------------------------------------------------------
# Logging Setup
# --------------------------------------------------------------
logger = logging.getLogger(__name__)
base_path = os.path.dirname(__file__)
log_dir = os.path.join(base_path, 'logs')
os.makedirs(log_dir, exist_ok=True)


# --------------------------------------------------------------
# Main Sweep Routine
# --------------------------------------------------------------
def run_sweep():
    """
    Run a complete frequency sweep test using the configured parameters.
    Performs baseline and selectively DPD-corrected measurements at each frequency,
    then saves results + summary statistics to Excel.
    """
    # -------------------------
    # File paths for test inputs, calibration data, and output
    # -------------------------
    json_path = os.path.join(base_path, 'test_inputs.json')
    cal_path = os.path.join(base_path, 'combined_cal_data.xlsx')
    output_path = os.path.join(base_path, 'sweep_measurements.xlsx')

    # -------------------------
    # Verify required files exist
    # -------------------------
    if not os.path.exists(json_path):
        logger.error(f"Test inputs file not found: {os.path.abspath(json_path)}")
        return
    if not os.path.exists(cal_path):
        logger.error(f"Calibration data file not found: {os.path.abspath(cal_path)}")
        return

    # -------------------------
    # Load sweep configuration from JSON
    # -------------------------
    with open(json_path, 'r') as f:
        config = json.load(f)

    sweep_measurement = config.get("Sweep_Measurement", {})
    sweep_params = sweep_measurement.get("range", {})
    signal_bandwidth = sweep_measurement.get("signal_bandwidth", "10MHz")
    frame_type = sweep_measurement.get("frame_type", "full_frame")
    user_comment_mode = sweep_measurement.get("user_comment_mode", "full_frame_nrx")

    # DPD enable flags
    enable_polynomial_dpd = sweep_measurement.get("enable_polynomial_dpd", True)
    enable_direct_dpd = sweep_measurement.get("enable_direct_dpd", True)
    enable_gmp_dpd = sweep_measurement.get("enable_gmp_dpd", True)

    # Envelope Tracking enable and parameters
    enable_et = sweep_measurement.get("enable_envelope_tracking", False)
    et_start_delay = sweep_measurement.get("et_starting_delay", 0.0)
    et_delay_shifts = sweep_measurement.get("et_delay_shifts", 0)
    et_delay_step = sweep_measurement.get("et_delay_step", 0.0)

    # Construct comment key based on signal_bandwidth and user_comment_mode
    comment_key = f"{signal_bandwidth}_{user_comment_mode}"
    comment_lines = config.get("User_Comments", {}).get(comment_key, [])
    if not comment_lines:
        logger.warning(f"No comments found for key '{comment_key}', using default")
        comment_key = f"{signal_bandwidth}_full_frame_nrx"
        comment_lines = config.get("User_Comments", {}).get(comment_key, [])
    user_comment = "\n".join(comment_lines)

    # Sweep setup
    start = sweep_params["start_ghz"] * 1e9
    stop = sweep_params["stop_ghz"] * 1e9
    step = sweep_params["step_mhz"] * 1e6
    target_output = sweep_params.get("power_dbm", 6.0)
    tolerance = sweep_params.get("tolerence_db", 0.05)
    expected_gain = sweep_params.get("expected_gain_db", 18.0)
    ddpd_iterations = sweep_params.get("ddpd_iterations", 5)
    servo_iterations = sweep_params.get("servo_iterations", 5)

    # Generate frequency list
    freqs = np.arange(start, stop + step, step)
    freqs_ghz = [round(f / 1e9, 3) for f in freqs]

    # -------------------------
    # Load calibration offsets (Excel → dictionary)
    # -------------------------
    cal_df = pd.read_excel(cal_path)
    cal_dict = {
        round(float(row["Center Frequency (GHz)"]), 3): {
            "vsg_offset": float(row["VSG Offset (dB)"]),
            "vsa_offset": float(row["VSA Offset (dB)"]),
            "input_offset": float(row["Input Power Offset (dB)"]),
            "output_offset": float(row["Output Power Offset (dB)"])
        } for _, row in cal_df.iterrows()
    }

    # -------------------------
    # Initialize instruments
    # -------------------------
    vsg = VSG()   # Signal generator
    vsa = VSA()   # Signal analyzer
    pm = PowerMeter()  # NRX power meter
    power_servo = PowerServo(vsg, pm, vsa, max_iterations=servo_iterations, tolerance=tolerance)

    # Initialize ET if enabled
    et = ET(vsg, et_delay_shifts=et_delay_shifts, vsa=vsa, pm=pm) if enable_et else None

    results = []
    matched_frequencies = []

    try:
        # ==========================================================
        # Loop through each frequency in the sweep
        # ==========================================================
        for freq in freqs:
            freq_ghz = round(freq / 1e9, 3)
            print(f"\n--- Testing Frequency: {freq_ghz} GHz ---")

            # Skip frequencies not found in calibration table
            if freq_ghz not in cal_dict:
                logger.warning(f"No calibration data for {freq_ghz} GHz — skipping")
                continue

            matched_frequencies.append(freq_ghz)
            offsets = cal_dict[freq_ghz]
            vsg_offset = offsets["vsg_offset"]
            vsa_offset = offsets["vsa_offset"]
            input_offset = offsets["input_offset"]
            output_offset = offsets["output_offset"]

            # -------------------------
            # Configure instruments with calibration offsets
            # -------------------------
            pm.configure(freq, input_offset, output_offset)
            vsg.configure(freq, target_output - expected_gain, vsg_offset)
            vsa.configure(freq, vsa_offset)

            # -------------------------
            # Baseline measurement (before DPD)
            # -------------------------
            corrected_input, corrected_output = pm.measure()
            start_time = time()
            freq_str = f"{freq:.0f}"

            (vsa_power, evm_value, evm_time, chan_pow, adj_chan_lower, adj_chan_upper, aclr_time, total_evm_time, servo_loops,
             ext_servo_time, k18_time, baseline_et_data) = vsa.measure_evm(
                freq_str, vsa_offset, target_output, servo_iterations, freq_ghz, expected_gain, power_servo, et=et
            )

            # -------------------------
            # Polynomial DPD measurement (formerly Single DPD)
            # -------------------------
            poly_et_data = None
            if enable_polynomial_dpd:
                (poly_power, poly_evm, poly_time,
                 poly_chan_pow, poly_adj_chan_lower,
                 poly_adj_chan_upper, poly_aclr_time,
                 poly_total_time, poly_servo_loops,
                 poly_ext_servo_time, poly_k18_time, poly_et_data) = vsa.perform_polynomial_dpd(
                    freq_str, vsa_offset, target_output, servo_iterations,
                    freq_ghz, expected_gain, power_servo, et=et
                )
            else:
                poly_power = poly_evm = poly_time = poly_chan_pow = poly_adj_chan_lower = poly_adj_chan_upper = None
                poly_aclr_time = poly_total_time = poly_servo_loops = poly_ext_servo_time = poly_k18_time = None

            # -------------------------
            # Direct DPD measurement (formerly Iterative DPD)
            # -------------------------
            direct_et_data = None
            if enable_direct_dpd:
                vsg.configure(freq, target_output - expected_gain, vsg_offset)
                (direct_power, direct_evm, direct_time,
                 direct_chan_pow, direct_adj_chan_lower, direct_adj_chan_upper,
                 direct_aclr_time, direct_total_time,
                 direct_servo_loops, direct_ext_servo_time, direct_k18_time, direct_et_data) = vsa.perform_direct_dpd(
                    freq_str, vsa_offset, target_output, ddpd_iterations,
                    servo_iterations, freq_ghz, expected_gain, power_servo, et=et
                )
            else:
                direct_power = direct_evm = direct_time = direct_chan_pow = direct_adj_chan_lower = direct_adj_chan_upper = None
                direct_aclr_time = direct_total_time = direct_servo_loops = direct_ext_servo_time = direct_k18_time = None

            # -------------------------
            # GMP DPD measurement
            # -------------------------
            gmp_et_data = None
            if enable_gmp_dpd:
                vsg.configure(freq, target_output - expected_gain, vsg_offset)
                (gmp_power, gmp_evm, gmp_time,
                 gmp_chan_pow, gmp_adj_chan_lower, gmp_adj_chan_upper,
                 gmp_aclr_time, gmp_total_time,
                 gmp_servo_loops, gmp_ext_servo_time, gmp_k18_time, gmp_et_data) = vsa.perform_gmp_dpd(
                    freq_str, vsa_offset, target_output, ddpd_iterations,
                    servo_iterations, freq_ghz, expected_gain, power_servo, et=et
                )
            else:
                gmp_power = gmp_evm = gmp_time = gmp_chan_pow = gmp_adj_chan_lower = gmp_adj_chan_upper = None
                gmp_aclr_time = gmp_total_time = gmp_servo_loops = gmp_ext_servo_time = gmp_k18_time = None

            # -------------------------
            # Collect and store results for this frequency
            # -------------------------
            elapsed = time() - start_time
            print(f"Total measurement time at {freq_ghz} GHz:, , , {elapsed:.3f}")

            result = {
                # Instrument setup times
                "VSG Setup Time (s)": vsg.setup_time,
                "VSA Setup Time (s)": vsa.setup_time,

                # Frequency & servo information
                "Center Frequency (GHz)": freq_ghz,
                "Target Output Power (dBm)": target_output,
                "Servo Iterations": servo_loops,
                "Servo Settle Time (s)": ext_servo_time,

                # Baseline measurements
                "Corrected Input Power (dBm)": corrected_input,
                "Corrected Output Power (dBm)": corrected_output,
                "VSA Output Power (dBm)": vsa_power,
                "EVM (dB)": evm_value,
                "EVM Measure Time (s)": evm_time,
                "Channel Power (dBm)": chan_pow,
                "Lower Adjacent ACLR (dB)": adj_chan_lower,
                "Upper Adjacent ACLR (dB)": adj_chan_upper,
                "ACLR Measure Time (s)": aclr_time,

                # Polynomial DPD results (formerly Single DPD)
                "Polynomial DPD Power (dBm)": poly_power,
                "Polynomial DPD EVM (dB)": poly_evm,
                "Polynomial DPD Measure Time (s)": poly_time,
                "Polynomial DPD Channel Power (dBm)": poly_chan_pow,
                "Polynomial DPD Lower Adjacent ACLR (dB)": poly_adj_chan_lower,
                "Polynomial DPD Upper Adjacent ACLR (dB)": poly_adj_chan_upper,
                "Polynomial DPD ACLR Measure Time (s)": poly_aclr_time,
                "Polynomial DPD Total Time (s)": poly_total_time,
                "Polynomial DPD Servo Loops": poly_servo_loops,
                "Polynomial DPD Ext Servo Time (s)": poly_ext_servo_time,
                "Polynomial DPD K18 Servo Time (s)": poly_k18_time,

                # Direct DPD results (formerly Iterative DPD)
                "Direct DPD Power (dBm)": direct_power,
                "Direct DPD EVM (dB)": direct_evm,
                "Direct DPD Measure Time (s)": direct_time,
                "Direct DPD Channel Power (dBm)": direct_chan_pow,
                "Direct DPD Lower Adjacent ACLR (dB)": direct_adj_chan_lower,
                "Direct DPD Upper Adjacent ACLR (dB)": direct_adj_chan_upper,
                "Direct DPD ACLR Measure Time (s)": direct_aclr_time,
                "Direct DPD Total Time (s)": direct_total_time,
                "Direct DPD Servo Loops": direct_servo_loops,
                "Direct DPD Ext Servo Time (s)": direct_ext_servo_time,
                "Direct DPD K18 Servo Time (s)": direct_k18_time,

                # GMP DPD results
                "GMP Power (dBm)": gmp_power,
                "GMP EVM (dB)": gmp_evm,
                "GMP Measure Time (s)": gmp_time,
                "GMP Channel Power (dBm)": gmp_chan_pow,
                "GMP Lower Adjacent ACLR (dB)": gmp_adj_chan_lower,
                "GMP Upper Adjacent ACLR (dB)": gmp_adj_chan_upper,
                "GMP ACLR Measure Time (s)": gmp_aclr_time,
                "GMP Total Time (s)": gmp_total_time,
                "GMP Servo Loops": gmp_servo_loops,
                "GMP Ext Servo Time (s)": gmp_ext_servo_time,
                "GMP K18 Servo Time (s)": gmp_k18_time,

                # Summary
                "Total Elapsed Time (s)": round(elapsed, 3),
                "Comment": user_comment
            }

            # Add ET configuration parameters
            if enable_et:
                result["ET Starting Delay (s)"] = et_start_delay
                result["ET Delay Step (s)"] = et_delay_step
                result["ET Number of Shifts"] = et_delay_shifts

                # Add baseline ET data
                if baseline_et_data:
                    result["Baseline ET Delays (s)"] = ", ".join(f"{d:.2e}" for d in baseline_et_data["delays"])
                    result["Baseline ET EVMs (dB)"] = ", ".join(f"{e:.2f}" for e in baseline_et_data["evms"])
                    result["Baseline ET Total Time (s)"] = round(baseline_et_data["total_time"], 3)

                # Add polynomial DPD ET data
                if poly_et_data:
                    result["Polynomial DPD ET Delays (s)"] = ", ".join(f"{d:.2e}" for d in poly_et_data["delays"])
                    result["Polynomial DPD ET EVMs (dB)"] = ", ".join(f"{e:.2f}" for e in poly_et_data["evms"])
                    result["Polynomial DPD ET Total Time (s)"] = round(poly_et_data["total_time"], 3)

                # Add direct DPD ET data
                if direct_et_data:
                    result["Direct DPD ET Delays (s)"] = ", ".join(f"{d:.2e}" for d in direct_et_data["delays"])
                    result["Direct DPD ET EVMs (dB)"] = ", ".join(f"{e:.2f}" for e in direct_et_data["evms"])
                    result["Direct DPD ET Total Time (s)"] = round(direct_et_data["total_time"], 3)

                # Add GMP ET data
                if gmp_et_data:
                    result["GMP ET Delays (s)"] = ", ".join(f"{d:.2e}" for d in gmp_et_data["delays"])
                    result["GMP ET EVMs (dB)"] = ", ".join(f"{e:.2f}" for e in gmp_et_data["evms"])
                    result["GMP ET Total Time (s)"] = round(gmp_et_data["total_time"], 3)

            results.append(result)
            print(user_comment)

    finally:
        # -------------------------
        # Always close instruments
        # -------------------------
        vsg.close()
        vsa.close()
        pm.close()

        # -------------------------
        # Save all results to Excel
        # -------------------------
        df = pd.DataFrame(results)
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Measurements sheet
            df.to_excel(writer, sheet_name='Measurements', index=False)
            worksheet = writer.sheets['Measurements']

            # Wrap text in "Comment" column and resize rows
            for col_idx, col_name in enumerate(df.columns, start=1):
                if col_name == "Comment":
                    for row_idx in range(2, len(df)+2):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.alignment = Alignment(wrap_text=True)
                        value = str(cell.value) if cell.value else ""
                        lines = value.count("\n") + 1
                        worksheet.row_dimensions[row_idx].height = 15 * lines

            # Statistics sheet (summary across all frequencies)
            if not df.empty:
                stats_rows = [("Number of Tests", len(df))]
                numeric_cols = [c for c in df.columns if pd.api.types.is_numeric_dtype(df[c])]
                for col in numeric_cols:
                    stats_rows.extend([
                        (f"{col} - Max", df[col].max()),
                        (f"{col} - Min", df[col].min()),
                        (f"{col} - Mean", df[col].mean())
                    ])
                stats_df = pd.DataFrame(stats_rows, columns=["Metric", "Value"])
                stats_df.to_excel(writer, sheet_name='Statistics', index=False)

                # Highlight means for readability
                yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                bold_font = Font(bold=True, size=12)
                worksheet_stats = writer.sheets['Statistics']
                for row_idx in range(2, len(stats_rows)+2):
                    metric_cell = worksheet_stats.cell(row=row_idx, column=1)
                    value_cell = worksheet_stats.cell(row=row_idx, column=2)
                    if "Mean" in str(metric_cell.value):
                        metric_cell.fill = yellow_fill
                        value_cell.fill = yellow_fill
                        value_cell.font = bold_font

        logger.info(f"Saved sweep results to: {os.path.abspath(output_path)}")
        logger.info(f"Processed frequencies (GHz): {matched_frequencies}")


# --------------------------------------------------------------
# Script Entry Point
# --------------------------------------------------------------
if __name__ == "__main__":
    run_sweep()